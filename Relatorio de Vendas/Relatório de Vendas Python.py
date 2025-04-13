# Importando as bibliotecas necessárias
import csv
from collections import defaultdict
from plyer import notification
import tkinter as tk
from tkinter import messagebox
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
import matplotlib.pyplot as plt
import seaborn as sns
import schedule
import time
import re

def exibir_menu():
    """
    Exibe o menu de opções e retorna a opção escolhida.
    """
    print("\nMenu:")
    print("1: Adicionar novo produto")
    print("2: Atualizar produto existente")
    print("3: Remover produto")
    print("4: Sair do programa")
    opcao = input("Escolha uma opção: ")
    return opcao

# Função para adicionar dados de vendas a um arquivo CSV
def adicionar_dados(arquivo_csv, arquivo_excel):
    """
    Adiciona ou atualiza dados nos arquivos CSV e Excel, evitando duplicatas no Excel.

    Args:
        arquivo_csv (str): Caminho para o arquivo CSV.
        arquivo_excel (str): Caminho para o arquivo Excel.
    """
    with open(arquivo_csv, 'a', newline='', encoding='utf-8') as csvfile:
        escritor = csv.writer(csvfile)
        try:
            workbook = load_workbook(arquivo_excel)
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Produto", "Quantidade", "Preço"])

        while True:
            produto = input("Digite o nome do produto (ou 'sair' para finalizar): ")
            if produto.lower() == 'sair':
                break

            if not produto.strip():  # Verifica se o nome do produto está vazio
                print("Formato inválido. O nome do produto não pode estar vazio.")
                continue  # Volta para o início do loop

            quantidade = input("Digite a quantidade: ")
            preco = input("Digite o preço: ")
            registro = (produto, int(quantidade), float(preco))

            # Verifica se o registro já existe no Excel
            registros_existentes = set()
            for row in sheet.iter_rows(min_row=2, values_only=True):
                registros_existentes.add(tuple(row))

            if registro in registros_existentes:
                # Atualiza o registro existente
                opcao = input(f"O registro '{registro}' já existe. Deseja atualizar (U) ou adicionar um novo (N)? (U/N): ")
                if opcao.lower() == 'u':
                    for row in sheet.iter_rows(min_row=2):
                        if tuple([cell.value for cell in row]) == registro:
                            row[1].value = quantidade
                            row[2].value = preco
                            break
                    print(f"Registro '{registro}' atualizado com sucesso.")
                    continue  # Volta para o início do loop

            # Adiciona um novo registro
            escritor.writerow(registro)
            sheet.append(registro)
            print(f"Registro '{registro}' adicionado com sucesso.")

        workbook.save(arquivo_excel)
        print("Dados adicionados/atualizados nos arquivos CSV e Excel com sucesso.")

def adicionar_produto(arquivo_csv, arquivo_excel):
    """
    Adiciona um novo produto aos arquivos CSV e Excel.
    """
    adicionar_dados(arquivo_csv, arquivo_excel)

def atualizar_produto(arquivo_csv, arquivo_excel):
    """
    Atualiza um ou mais produtos existentes nos arquivos CSV e Excel.
    """
    # Exibe a lista de produtos
    print("Lista de produtos:")
    with open(arquivo_csv, 'r', newline='', encoding='utf-8') as arquivo:
        leitor_csv = csv.reader(arquivo)
        cabecalho = next(leitor_csv)  # Lê o cabeçalho
        for linha in leitor_csv:
            print(linha[0])  # Exibe o nome do produto

    while True:
        nome_produto = input("Digite o nome do produto que deseja atualizar (ou 'sair' para finalizar): ")
        if nome_produto.lower() == 'sair':
            break

        # Verifica se o produto existe
        produto_existe = False
        linhas = []
        with open(arquivo_csv, 'r', newline='', encoding='utf-8') as arquivo:
            leitor_csv = csv.reader(arquivo)
            cabecalho = next(leitor_csv)
            linhas.append(cabecalho)
            for linha in leitor_csv:
                if linha[0] == nome_produto:
                    produto_existe = True
                    quantidade = input("Digite a nova quantidade: ")
                    preco_unitario = input("Digite o novo preço unitário: ")
                    linha[1] = quantidade
                    linha[2] = preco_unitario
                linhas.append(linha)

        if not produto_existe:
            print("Produto não encontrado.")
            continue

        # Atualiza o arquivo CSV
        with open(arquivo_csv, 'w', newline='', encoding='utf-8') as arquivo:
            escritor_csv = csv.writer(arquivo)
            escritor_csv.writerows(linhas)

        # Atualiza o arquivo Excel
        df = pd.read_csv(arquivo_csv)
        df.to_excel(arquivo_excel, index=False)

        print("Produto atualizado com sucesso!")

def remover_produto(arquivo_csv, arquivo_excel):
    """
    Remove um ou mais produtos dos arquivos CSV e Excel e exibe os dados existentes.
    """
    try:
        df = pd.read_csv(arquivo_csv)
        print("\nDados existentes:")
        print(df)
    except FileNotFoundError:
        print(f"Erro: Arquivo {arquivo_csv} não encontrado.")
        return

    while True:
        produto = input("Digite o nome do produto a ser removido (ou 'sair' para finalizar): ")
        if produto.lower() == 'sair':
            break

        excluir_dados(arquivo_csv, produto)
        excluir_dados(arquivo_excel, produto)
        print(f"Produto '{produto}' removido com sucesso.")

# Função para ler os dados do CSV e calcular o total das vendas por produto
def ler_csv(arquivo):
    """
    Lê os dados do arquivo CSV e calcula o total de vendas por produto, considerando quantidade e preço.

    Args:
        arquivo (str): Caminho para o arquivo CSV.

    Returns:
        dict: Dicionário com o total de vendas por produto.
    """
    vendas = defaultdict(float)
    with open(arquivo, 'r', newline='', encoding='utf-8') as csvfile:
        leitor = csv.reader(csvfile)
        next(leitor)  # Pula o cabeçalho
        for linha in leitor:
            if linha:  # Verifica se a linha não está vazia
                produto, quantidade, preco = linha
                vendas[produto] += int(quantidade) * float(preco)
    return vendas

# Função para gerar um relatório simples das vendas
def gerar_relatorio(vendas):
    if not vendas:  # Verifica se não há vendas
        print("Nenhum dado de vendas disponível para gerar relatório.")
        return

    # Encontra o produto mais vendido e calcula o total geral
    produto_mais_vendido = max(vendas, key=vendas.get)
    total_geral = sum(vendas.values())

    print("Relatório de Vendas:")
    # Exibe o relatório no console
    for produto, total in vendas.items():
        print(f"{produto}: R$ {total:.2f}")
    print(f"\nProduto mais vendido: {produto_mais_vendido} (R$ {vendas[produto_mais_vendido]:.2f})")
    print(f"Total geral de vendas: R$ {total_geral:.2f}")

    # Chama funções para enviar notificações e exibir pop-ups
    enviar_notificacao(produto_mais_vendido, total_geral)
    exibir_popup(produto_mais_vendido, total_geral)

# Função para enviar uma notificação de desktop
def enviar_notificacao(produto_mais_vendido, total_geral):
    mensagem = f"Produto mais vendido: {produto_mais_vendido}\nTotal de vendas: R$ {total_geral:.2f}"
    notification.notify(
        title="Resumo de Vendas",  # Título da notificação
        message=mensagem,  # Mensagem da notificação
        app_name="Sistema de Vendas"  # Nome do aplicativo
    )

# Função para exibir um pop-up com o resumo das vendas
def exibir_popup(produto_mais_vendido, total_geral):
    """
    Exibe um pop-up com o resumo das vendas, sempre em primeiro plano e com tamanho ajustado.
    """
    root = tk.Tk()
    root.title("Resumo de Vendas")  # Título completo
    root.attributes('-topmost', True)

    mensagem = f"Produto mais vendido: {produto_mais_vendido}\nTotal de vendas: R$ {total_geral:.2f}"
    label = tk.Label(root, text=mensagem, font=("Arial", 14))  # Aumenta a fonte
    label.pack(padx=30, pady=30)  # Aumenta o padding

    botao_ok = tk.Button(root, text="OK", command=root.destroy, font=("Arial", 12))  # Aumenta a fonte do botão
    botao_ok.pack(pady=15)

    root.mainloop()

# Função para gerar um relatório de vendas em formato Excel
from openpyxl import load_workbook

from openpyxl import load_workbook

from openpyxl import load_workbook

def gerar_relatorio_excel(arquivo_csv, arquivo_excel):
    """
    Lê um arquivo CSV de vendas, verifica duplicatas e adiciona apenas novos registros na planilha existente.

    Args:
        arquivo_csv (str): Caminho para o arquivo CSV de vendas.
        arquivo_excel (str): Caminho para o arquivo Excel de saída.
    """
    try:
        # Ler o arquivo CSV usando pandas
        df_novos = pd.read_csv(arquivo_csv)

        # Verifica se o arquivo Excel já existe
        if os.path.exists(arquivo_excel):
            workbook = load_workbook(arquivo_excel)
            if "Resumo" in workbook.sheetnames:
                sheet = workbook["Resumo"]
            else:
                sheet = workbook.create_sheet(title="Resumo")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Resumo"

        # Criar uma estrutura para armazenar registros existentes
        registros_existentes = set()

        # Se a planilha não estiver vazia, carregar os registros atuais para evitar duplicação
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Começa da linha 2 para ignorar cabeçalho
            registros_existentes.add(tuple(row))

        # Descobrir a próxima linha disponível corretamente
        ultima_linha = sheet.max_row

        # Se a planilha estiver vazia, adicionar cabeçalho na primeira linha
        if ultima_linha == 1 and sheet.cell(row=1, column=1).value is None:
            sheet.append(["Produto", "Quantidade", "Preço", "Total"])
            ultima_linha = 1

        # Adiciona apenas os novos dados que não estão na planilha
        for _, row in df_novos.iterrows():
            produto = row["Produto"]
            quantidade = row["Quantidade"]
            preco = row["Preço"]
            total = quantidade * preco

            novo_registro = (produto, quantidade, preco, total)

            if novo_registro not in registros_existentes:  # Evita duplicação
                sheet.append(novo_registro)

        # Salva o arquivo Excel atualizado
        workbook.save(arquivo_excel)

        print(f"Relatório atualizado sem duplicatas em {arquivo_excel}")

    except FileNotFoundError:
        print(f"Erro: Arquivo {arquivo_csv} não encontrado.")
    except Exception as e:
        print(f"Ocorreu um erro: {e}")

def excluir_dados(arquivo, produto_excluir):
    """
    Exclui linhas do arquivo CSV ou Excel com base no nome do produto.

    Args:
        arquivo (str): Caminho para o arquivo (CSV ou Excel).
        produto_excluir (str): Nome do produto a ser excluído.
    """
    if arquivo.lower().endswith('.csv'):
        try:
            with open(arquivo, 'r', newline='', encoding='utf-8') as csvfile:
                leitor = csv.reader(csvfile)
                linhas = [linha for linha in leitor if linha and linha[0].lower() != produto_excluir.lower()]

            with open(arquivo, 'w', newline='', encoding='utf-8') as csvfile:
                escritor = csv.writer(csvfile)
                escritor.writerows(linhas)

            print(f"Dados do produto '{produto_excluir}' excluídos do CSV com sucesso.")

        except FileNotFoundError:
            print(f"Erro: Arquivo {arquivo} não encontrado.")
        except Exception as e:
            print(f"Ocorreu um erro ao excluir dados do CSV: {e}")

    elif arquivo.lower().endswith('.xlsx'):
        try:
            workbook = load_workbook(arquivo)
            sheet = workbook.active
            linhas_para_excluir = []

            for linha_num, linha in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if linha and linha[0].lower() == produto_excluir.lower():
                    linhas_para_excluir.append(linha_num)

            for linha_num in reversed(linhas_para_excluir):
                sheet.delete_rows(linha_num)

            workbook.save(arquivo)
            print(f"Dados do produto '{produto_excluir}' excluídos do Excel com sucesso.")

        except FileNotFoundError:
            print(f"Erro: Arquivo {arquivo} não encontrado.")
        except Exception as e:
            print(f"Ocorreu um erro ao excluir dados do Excel: {e}")
    else:
        print("Formato de arquivo não suportado. Use CSV ou XLSX.")

def gerar_graficos(arquivo_csv, arquivo_grafico_quantidade, arquivo_grafico_preco, arquivo_grafico_combinado):
    """
    Gera gráficos de quantidade de vendas, preço unitário e um gráfico combinado.

    Args:
        arquivo_csv (str): Caminho para o arquivo CSV.
        arquivo_grafico_quantidade (str): Caminho para salvar o gráfico de quantidade.
        arquivo_grafico_preco (str): Caminho para salvar o gráfico de preço.
        arquivo_grafico_combinado (str): Caminho para salvar o gráfico combinado.
    """
    df = pd.read_csv(arquivo_csv)

    # Gráfico de quantidade de vendas
    plt.figure(figsize=(10, 6))
    sns.barplot(x='Produto', y='Quantidade', data=df)
    plt.title('Quantidade de Vendas por Produto')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(arquivo_grafico_quantidade)
    plt.close()

    # Gráfico de preço unitário
    plt.figure(figsize=(10, 6))
    sns.barplot(x='Produto', y='Preço', data=df)
    plt.title('Preço Unitário por Produto')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(arquivo_grafico_preco)
    plt.close()

    # Gráfico combinado (quantidade e preço)
    plt.figure(figsize=(12, 7))
    ax1 = plt.gca()
    sns.barplot(x='Produto', y='Quantidade', data=df, ax=ax1, color='skyblue')
    ax2 = ax1.twinx()
    sns.lineplot(x='Produto', y='Preço', data=df, ax=ax2, color='red', marker='o')
    plt.title('Quantidade de Vendas e Preço Unitário por Produto')
    ax1.set_ylabel('Quantidade de Vendas', color='skyblue')
    ax2.set_ylabel('Preço Unitário', color='red')
    plt.xticks(rotation=45, ha='right')
    plt.tight_layout()
    plt.savefig(arquivo_grafico_combinado)
    plt.close()

def enviar_email(arquivo_excel, arquivo_grafico_quantidade, arquivo_grafico_preco, arquivo_grafico_combinado, destinatario, remetente, senha):
    """
    Envia um e-mail com o relatório Excel e os gráficos anexados.

    Args:
        arquivo_excel (str): Caminho para o arquivo Excel do relatório.
        arquivo_grafico_quantidade (str): Caminho para o gráfico de quantidade.
        arquivo_grafico_preco (str): Caminho para o gráfico de preço.
        arquivo_grafico_combinado (str): Caminho para o gráfico combinado.
        destinatario (str): Endereço de e-mail do destinatário.
        remetente (str): Endereço de e-mail do remetente.
        senha (str): Senha do remetente.
    """
    try:
        msg = MIMEMultipart()
        msg['From'] = remetente
        msg['To'] = destinatario
        msg['Subject'] = "Relatório de Vendas com Gráficos"

        body = "Segue em anexo o relatório de vendas e os gráficos."
        msg.attach(MIMEText(body, 'plain'))

        # Anexa os arquivos
        anexar_arquivo(msg, arquivo_excel, 'xlsx')
        anexar_arquivo(msg, arquivo_grafico_quantidade, 'png')
        anexar_arquivo(msg, arquivo_grafico_preco, 'png')
        anexar_arquivo(msg, arquivo_grafico_combinado, 'png')

        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
            server.login(remetente, senha)
            server.send_message(msg)

        print("E-mail enviado com sucesso!")

    except Exception as e:
        print(f"Ocorreu um erro ao enviar o e-mail: {e}")

def anexar_arquivo(msg, arquivo, subtype):
    with open(arquivo, "rb") as f:
        attach = MIMEApplication(f.read(), _subtype=subtype)
        attach.add_header('Content-Disposition', 'attachment', filename=os.path.basename(arquivo))
        msg.attach(attach)

# Função para obter as credenciais de e-mail do usuário através de uma interface gráfica
def obter_credenciais_email():
    """
    Coleta as credenciais de e-mail do usuário através de uma interface gráfica Tkinter.
    """
    root = tk.Tk()
    root.title("Credenciais de E-mail")

    root.geometry("400x250")
    root.attributes('-topmost', True)

    remetente_label = tk.Label(root, text="E-mail do Remetente:")
    remetente_label.pack()
    remetente_entry = tk.Entry(root, width=40)
    remetente_entry.pack()

    senha_label = tk.Label(root, text="Senha:")
    senha_label.pack()
    senha_entry = tk.Entry(root, show="*", width=40)
    senha_entry.pack()

    destinatario_label = tk.Label(root, text="E-mail do Destinatário:")
    destinatario_label.pack()
    destinatario_entry = tk.Entry(root, width=40)
    destinatario_entry.pack()

    def enviar_e_destruir():  # Função interna
        root.credenciais = enviar()
        if root.credenciais:
            root.destroy()

    def enviar():
        remetente = remetente_entry.get()
        senha = senha_entry.get()
        destinatario = destinatario_entry.get()

        if not validar_email(remetente):
            messagebox.showerror("Erro", "E-mail do remetente inválido")
            return None

        if not validar_email(destinatario):
            messagebox.showerror("Erro", "E-mail do destinatário inválido")
            return None

        return remetente, senha, destinatario

    def validar_email(email):
        padrao = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$"
        if re.match(padrao, email):
            return True
        else:
            return False

    enviar_button = tk.Button(root, text="Enviar", command=enviar_e_destruir)  # Usando função interna
    enviar_button.pack()

    root.mainloop()

    if hasattr(root, 'credenciais'):
        return root.credenciais
    else:
        return None

# Código principal
if __name__ == "__main__":
    pasta_raiz = r"C:\Users\adria\Videos\Citha-Python\10 exercicios"
    arquivo_csv = os.path.join(pasta_raiz, "vendas.csv")
    arquivo_excel = os.path.join(pasta_raiz, "relatorio.xlsx")
    arquivo_grafico_quantidade = os.path.join(pasta_raiz, "grafico_quantidade.png")
    arquivo_grafico_preco = os.path.join(pasta_raiz, "grafico_preco.png")
    arquivo_grafico_combinado = os.path.join(pasta_raiz, "grafico_combinado.png")

    while True:
        opcao = exibir_menu()

        if opcao == '1':
            adicionar_produto(arquivo_csv, arquivo_excel)
        elif opcao == '2':
            atualizar_produto(arquivo_csv, arquivo_excel)
        elif opcao == '3':
            remover_produto(arquivo_csv, arquivo_excel)
        elif opcao == '4':
            break
        else:
            print("Opção inválida. Tente novamente.")
            continue

        vendas = ler_csv(arquivo_csv)
        gerar_relatorio(vendas)
        gerar_relatorio_excel(arquivo_csv, arquivo_excel)
        gerar_graficos(arquivo_csv, arquivo_grafico_quantidade, arquivo_grafico_preco, arquivo_grafico_combinado)

        credenciais = obter_credenciais_email()
        if credenciais:
            remetente, senha, destinatario = credenciais
            enviar_email(arquivo_excel, arquivo_grafico_quantidade, arquivo_grafico_preco, arquivo_grafico_combinado, destinatario, remetente, senha)